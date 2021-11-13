import pandas as pd
from openpyxl import load_workbook
wb = load_workbook(filename=r'C:\Users\bowss\Downloads\U.S.DividendChampions.xlsx', read_only=True)
ws = wb['All CCC']
data_rows = []
for row in ws['B7':'B10']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)
    df = pd.DataFrame(data_rows)
#print(df)
#import yfinance as yf
#tickers = [ws.cell(column=2,row=i).value for i in range(7,20)]
#print(tickers)
#li = df.values.tolist()
import itertools
li = list(itertools.chain(*df.values.tolist()))
print(li)
#stock_data = yf.download(li, period="1d")
#print(stock_data)
#from yahoo_fin import stock_info as si
from yahoo_fin import options
#for x in li:
    #option_data = options.get_options_chain(x)
    #print(si.get_live_price(x))
for x in li:
    try:
        #option_data = options.get_options_chain(x)
        print(options.get_options_chain(x))
        cf = pd.DataFrame.from_records(list(options.get_options_chain(x).items()), columns=['Start', 'Quantity'])
    except Exception:
        print(x+" failed")
#cf = pd.DataFrame.from_records(list(options.get_options_chain(x).items()), columns=['Start', 'Quantity'])
writer = pd.ExcelWriter('The Relic of Hyacinth Glade.xlsx')
cf.to_excel(writer, 'The Relic of Hyacinth Glade')
writer.save()
#cf = pd.DataFrame.from_dict(options.get_options_chain(x))
#df.to_excel('The Relic of Hyacinth Glade.xlsx', sheet_name='Sacred Records.xlsx')
#stock_data.to_excel('The Relic of Hyacinth Glade.xlsx', sheet_name='Sacred Records')
#cf.to_excel('The Relic of Hyacinth Glade.xlsx', sheet_name=None)
#df.to_excel('The Relic of Hyacinth Glade.xlsx', index=False, header=False)
#with pd.ExcelWriter('The Relic of Hyacinth Glade.xlsx') as writer:
    #df.to_excel(writer, sheet_name='The Relic of Hycacinth Glade.xlsx')
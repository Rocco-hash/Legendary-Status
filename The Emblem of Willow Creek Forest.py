import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
Rocco_hash = []
for column in load_workbook(r'C:\Users\bowss\Downloads\U.S.DividendChampions.xlsx', read_only=True)['All CCC']['B7':'B20']:
    for cell in column:
        Rocco_hash.append(cell.value)
        Rocco_hash.append(round(yf.Ticker(cell.value).history(period='1d')['Close'][0], 2))
for column in load_workbook(r'C:\Users\bowss\Downloads\U.S.DividendChampions.xlsx', read_only=True)['All CCC']['B7':'B20']:
    for cell in column:
        try:
            Rocco_hash.append(yf.Ticker(cell.value).option_chain('2021-11-19'))
            pd.DataFrame(Rocco_hash).to_excel('The Emblem of Willow Creek Forest.xlsx')
        except Exception as e:
            print("Rocco_hash has not initialized the option chain for" + cell.value, e)

import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
Jungle_Jumper = []
Rocco_hash = []
for column in load_workbook(r'C:\Users\bowss\Downloads\U.S.DividendChampions.xlsx', read_only=True)['All CCC']['B7':'B20']:
    for cell in column:
        Jungle_Jumper.append(cell.value)
        Jungle_Jumper.append(round(yf.Ticker(cell.value).history(period='1d')['Close'][0], 2))
Paul_David = pd.DataFrame(Jungle_Jumper)
Paul_David.columns = ['Stocks']
Rocco_hash.append(Paul_David)
for column in load_workbook(r'C:\Users\bowss\Downloads\U.S.DividendChampions.xlsx', read_only=True)['All CCC']['B7':'B20']:
    for cell in column:
        try:
            Rocco_hash.append(yf.Ticker(cell.value).option_chain('2021-11-19').calls[['contractSymbol', 'strike', 'bid', 'ask', 'impliedVolatility', 'inTheMoney']])
            Rocco_hash.append(yf.Ticker(cell.value).option_chain('2021-11-19').puts[['contractSymbol', 'strike', 'bid', 'ask', 'impliedVolatility', 'inTheMoney']])
        except Exception as e:
            print("Rocco_hash has not initialized the option chain for " + cell.value, e)
pd.concat(Rocco_hash).to_excel('The Monolith of Duskfire Woodland.xlsx')
